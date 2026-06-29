#!/usr/bin/env python3
"""
LATAM Airlines CGH — Gerador de data.js
========================================
Script para processar a planilha Excel e gerar o arquivo js/data.js
usado pelo dashboard HTML.

Uso:
    python gerar_data.py
    python gerar_data.py --arquivo "caminho/para/planilha.xlsx"
    python gerar_data.py --aba "RELAÇÃO GERAL COLABORADORES CGH"

Dependências:
    pip install openpyxl
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from datetime import date, datetime


# ── CONFIGURAÇÕES PADRÃO ─────────────────────────────────────────────────────

ARQUIVO_PADRAO = 'CGH_ATUALIZADA_COM_EMAILS_COMPLETO.xlsx'
ABA_PADRAO     = 'RELAÇÃO GERAL COLABORADORES CGH'
SAIDA_PADRAO   = os.path.join(os.path.dirname(__file__), '..', 'js', 'data.js')

# Índices das colunas (0-based)
COL = {
    'mes':           0,
    'bp':            1,
    'nome':          2,
    'email':         3,
    'cargo':         4,
    'base':          5,
    'diretoria':     6,
    'email_gestor':  7,
    'gestor':        8,
    'ativo':         9,
    'regional':      10,
    'curso':         11,
    'sigla_curso':   12,
    'categoria':     13,
    'modalidade':    14,
    'data_real':     15,
    'validade':      16,
    'vencimento':    17,
    'status':        18,
    'emoji':         19,
    'prox_venc':     20,
    'area':          21,
}


# ── PROCESSAMENTO ─────────────────────────────────────────────────────────────

def parse_date(val):
    """Tenta converter valor para date. Retorna None se não for possível."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def determinar_alerta(dias):
    """Retorna o tipo de alerta com base nos dias para vencer."""
    if dias is None:
        return None
    if dias < 0:
        return 'VENCIDO'
    if dias <= 7:
        return '7_DIAS'
    if dias <= 15:
        return '15_DIAS'
    if dias <= 30:
        return '30_DIAS'
    if dias <= 60:
        return '60_DIAS'
    return None


def processar_planilha(arquivo, aba):
    """Lê a planilha e retorna lista de registros processados."""
    try:
        import openpyxl
    except ImportError:
        print('ERRO: openpyxl não instalado. Execute: pip install openpyxl')
        sys.exit(1)

    print(f'📂 Lendo arquivo: {arquivo}')
    print(f'📋 Aba: {aba}')

    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)

    if aba not in wb.sheetnames:
        print(f'ERRO: Aba "{aba}" não encontrada.')
        print(f'Abas disponíveis: {", ".join(wb.sheetnames)}')
        sys.exit(1)

    ws = wb[aba]
    registros = []
    headers_lido = False

    for row in ws.iter_rows(values_only=True):
        if not headers_lido:
            headers_lido = True
            continue  # pula cabeçalho

        # Extrai campos
        bp     = str(row[COL['bp']])    if row[COL['bp']]    else None
        nome   = str(row[COL['nome']])  if row[COL['nome']]  else None
        ativo  = str(row[COL['ativo']]) if row[COL['ativo']] else ''
        curso  = str(row[COL['curso']]) if row[COL['curso']] else None

        if not bp or not nome or not curso:
            continue
        if 'DESLIGAD' in ativo.upper():
            continue

        email        = str(row[COL['email']])        if row[COL['email']]        else ''
        cargo        = str(row[COL['cargo']])        if row[COL['cargo']]        else ''
        gestor       = str(row[COL['gestor']])       if row[COL['gestor']]       else 'Sem Gestor'
        email_gestor = str(row[COL['email_gestor']]) if row[COL['email_gestor']] else ''
        area         = str(row[COL['area']])         if row[COL['area']]         else 'ATO'
        status       = str(row[COL['status']])       if row[COL['status']]       else 'PENDENTE'
        prox_venc    = str(row[COL['prox_venc']])    if row[COL['prox_venc']]    else 'NÃO'

        # Vencimento
        venc_date = parse_date(row[COL['vencimento']])
        venc_str  = venc_date.strftime('%Y-%m-%d') if venc_date else ''

        registros.append({
            'bp':            bp,
            'nome':          nome,
            'email':         email,
            'cargo':         cargo,
            'gestor':        gestor,
            'email_gestor':  email_gestor,
            'area':          area if area in ('ATO', 'GRH') else 'ATO',
            'curso':         curso,
            'status':        status,
            'vencimento':    venc_str,
            'prox_venc':     prox_venc,
        })

    print(f'✅ {len(registros)} registros lidos')
    return registros


def calcular_stats(registros):
    """Calcula todas as estatísticas necessárias para o dashboard."""
    hoje = date.today()

    # Índice de colaboradores
    colabs = {}
    for r in registros:
        bp = r['bp']
        if bp not in colabs:
            colabs[bp] = {
                'bp':           bp,
                'nome':         r['nome'],
                'email':        r['email'],
                'cargo':        r['cargo'],
                'gestor':       r['gestor'],
                'email_gestor': r['email_gestor'],
                'area':         r['area'],
                'cursos':       []
            }

        venc_date = None
        if r['vencimento']:
            try:
                venc_date = datetime.strptime(r['vencimento'], '%Y-%m-%d').date()
            except ValueError:
                pass

        dias = (venc_date - hoje).days if venc_date else None
        alerta = determinar_alerta(dias)

        colabs[bp]['cursos'].append({
            'curso':           r['curso'],
            'status':          r['status'],
            'vencimento_date': r['vencimento'],
            'dias_para_vencer': dias,
            'alerta':          alerta,
            'area':            r['area'],
        })

    # ── Summary ──────────────────────────────────────────────
    total_ok   = sum(1 for r in registros if r['status'].upper() == 'OK')
    total_pend = sum(1 for r in registros if r['status'].upper() == 'PENDENTE')

    alertas_count = {'VENCIDO': 0, '7_DIAS': 0, '15_DIAS': 0, '30_DIAS': 0, '60_DIAS': 0}
    for c in colabs.values():
        for cr in c['cursos']:
            al = cr['alerta']
            if al and al in alertas_count:
                alertas_count[al] += 1

    conformidade = round(total_ok / (total_ok + total_pend) * 100, 1) if (total_ok + total_pend) else 0

    # Colaboradores com pendência
    colabs_pend = []
    for bp, c in colabs.items():
        pend_list = [cr for cr in c['cursos'] if cr['status'].upper() == 'PENDENTE']
        if pend_list:
            colabs_pend.append({
                'bp':          c['bp'],
                'nome':        c['nome'],
                'email':       c['email'],
                'cargo':       c['cargo'],
                'gestor':      c['gestor'],
                'email_gestor':c['email_gestor'],
                'area':        c['area'],
                'pendentes':   pend_list,
                'qtd_pend':    len(pend_list)
            })
    colabs_pend.sort(key=lambda x: -x['qtd_pend'])

    summary = {
        'total_colabs':        len(colabs),
        'total_registros':     len(registros),
        'data_atualizacao':    hoje.strftime('%d/%m/%Y'),
        'total_ok':            total_ok,
        'total_pend':          total_pend,
        'conformidade':        conformidade,
        'vencidos':            alertas_count['VENCIDO'],
        'vence_7':             alertas_count['7_DIAS'],
        'vence_15':            alertas_count['15_DIAS'],
        'vence_30':            alertas_count['30_DIAS'],
        'vence_60':            alertas_count['60_DIAS'],
        'colabs_com_pendencia': len(colabs_pend),
    }

    # ── Gestor stats ─────────────────────────────────────────
    gestor_stats = defaultdict(lambda: {'ok': 0, 'pend': 0, 'pessoas': set(), 'email': ''})
    for bp, c in colabs.items():
        g = c['gestor'] or 'Sem Gestor'
        gestor_stats[g]['pessoas'].add(bp)
        gestor_stats[g]['email'] = c['email_gestor']
        for cr in c['cursos']:
            if cr['status'].upper() == 'OK':
                gestor_stats[g]['ok'] += 1
            elif cr['status'].upper() == 'PENDENTE':
                gestor_stats[g]['pend'] += 1

    gestor_final = {
        g: {'ok': v['ok'], 'pend': v['pend'], 'pessoas': len(v['pessoas']), 'email': v['email']}
        for g, v in gestor_stats.items()
    }

    # ── Cargo stats ───────────────────────────────────────────
    cargo_stats = defaultdict(lambda: {'ok': 0, 'pend': 0, 'pessoas': set()})
    for bp, c in colabs.items():
        cargo = c['cargo'] or 'Sem Cargo'
        cargo_stats[cargo]['pessoas'].add(bp)
        for cr in c['cursos']:
            if cr['status'].upper() == 'OK':
                cargo_stats[cargo]['ok'] += 1
            elif cr['status'].upper() == 'PENDENTE':
                cargo_stats[cargo]['pend'] += 1

    cargo_final = {
        cg: {'ok': v['ok'], 'pend': v['pend'], 'pessoas': len(v['pessoas'])}
        for cg, v in cargo_stats.items()
    }

    # ── Curso stats ───────────────────────────────────────────
    curso_stats = defaultdict(lambda: {'ok': 0, 'pend': 0})
    for r in registros:
        if r['status'].upper() == 'OK':
            curso_stats[r['curso']]['ok'] += 1
        elif r['status'].upper() == 'PENDENTE':
            curso_stats[r['curso']]['pend'] += 1

    # ── Area stats ────────────────────────────────────────────
    area_stats = {
        'ATO': {'ok': 0, 'pend': 0, 'pessoas': set()},
        'GRH': {'ok': 0, 'pend': 0, 'pessoas': set()}
    }
    for r in registros:
        area = r['area'] if r['area'] in ('ATO', 'GRH') else 'ATO'
        area_stats[area]['pessoas'].add(r['bp'])
        if r['status'].upper() == 'OK':
            area_stats[area]['ok'] += 1
        elif r['status'].upper() == 'PENDENTE':
            area_stats[area]['pend'] += 1

    area_final = {
        k: {'ok': v['ok'], 'pend': v['pend'], 'pessoas': len(v['pessoas'])}
        for k, v in area_stats.items()
    }

    return {
        'summary':       summary,
        'colab_pend':    colabs_pend[:300],
        'gestor_stats':  gestor_final,
        'cargo_stats':   cargo_final,
        'curso_stats':   dict(curso_stats),
        'area_stats':    area_final,
        'all_colabs':    list(colabs.values())[:1200],
    }


def gerar_data_js(stats, caminho_saida):
    """Gera o arquivo data.js com todas as estatísticas."""
    s = stats['summary']

    js = f"""// ============================================================
// LATAM Airlines CGH — Dados do Dashboard
// Gerado automaticamente em: {s['data_atualizacao']}
// NÃO EDITE MANUALMENTE — use scripts/gerar_data.py
// ============================================================

window.LATAM_DATA = {{

  // ── RESUMO EXECUTIVO ─────────────────────────────────────
  summary: {json.dumps(s, ensure_ascii=False, indent=2)},

  // ── COLABORADORES COM PENDÊNCIAS ─────────────────────────
  colab_pend: {json.dumps(stats['colab_pend'], ensure_ascii=False)},

  // ── ESTATÍSTICAS POR GESTOR ──────────────────────────────
  gestor_stats: {json.dumps(stats['gestor_stats'], ensure_ascii=False)},

  // ── ESTATÍSTICAS POR CARGO ───────────────────────────────
  cargo_stats: {json.dumps(stats['cargo_stats'], ensure_ascii=False)},

  // ── ESTATÍSTICAS POR CURSO ───────────────────────────────
  curso_stats: {json.dumps(stats['curso_stats'], ensure_ascii=False)},

  // ── ESTATÍSTICAS POR ÁREA (ATO / GRH) ───────────────────
  area_stats: {json.dumps(stats['area_stats'], ensure_ascii=False)},

  // ── TODOS OS COLABORADORES (busca) ───────────────────────
  all_colabs: {json.dumps(stats['all_colabs'], ensure_ascii=False)}

}};
"""

    # Garante diretório de saída
    os.makedirs(os.path.dirname(os.path.abspath(caminho_saida)), exist_ok=True)

    with open(caminho_saida, 'w', encoding='utf-8') as f:
        f.write(js)

    tamanho_kb = os.path.getsize(caminho_saida) // 1024
    print(f'✅ data.js gerado: {caminho_saida} ({tamanho_kb} KB)')


def imprimir_resumo(stats):
    """Exibe resumo no terminal."""
    s = stats['summary']
    print()
    print('=' * 50)
    print('  LATAM CGH — Resumo do Processamento')
    print('=' * 50)
    print(f"  Colaboradores:       {s['total_colabs']:>8,}")
    print(f"  Registros:           {s['total_registros']:>8,}")
    print(f"  Cursos em dia:       {s['total_ok']:>8,}")
    print(f"  Pendências:          {s['total_pend']:>8,}")
    print(f"  Conformidade:        {s['conformidade']:>7}%")
    print(f"  Colabs c/ pendência: {s['colabs_com_pendencia']:>8,}")
    print(f"  Vencidos:            {s['vencidos']:>8,}")
    print(f"  Vence em 7 dias:     {s['vence_7']:>8,}")
    print(f"  Vence em 30 dias:    {s['vence_30']:>8,}")
    print(f"  Vence em 60 dias:    {s['vence_60']:>8,}")
    print('=' * 50)
    print()


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Gera js/data.js a partir da planilha Excel LATAM CGH'
    )
    parser.add_argument('--arquivo', default=ARQUIVO_PADRAO,
                        help=f'Caminho do arquivo Excel (padrão: {ARQUIVO_PADRAO})')
    parser.add_argument('--aba', default=ABA_PADRAO,
                        help=f'Nome da aba (padrão: {ABA_PADRAO})')
    parser.add_argument('--saida', default=SAIDA_PADRAO,
                        help=f'Caminho do arquivo de saída (padrão: {SAIDA_PADRAO})')
    args = parser.parse_args()

    if not os.path.exists(args.arquivo):
        print(f'ERRO: Arquivo não encontrado: {args.arquivo}')
        sys.exit(1)

    registros = processar_planilha(args.arquivo, args.aba)
    stats = calcular_stats(registros)
    gerar_data_js(stats, args.saida)
    imprimir_resumo(stats)
    print('🚀 Pronto! Faça o commit de js/data.js para atualizar o dashboard.')


if __name__ == '__main__':
    main()
